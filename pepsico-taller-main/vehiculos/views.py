
# Create your views here.
# vehiculos/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count
from django.db.models import Prefetch
from .models import Vehiculo, DocumentoVehiculo, Siniestro, FotoSiniestro  
from inventario.models import Repuesto
from mantenimientos.models import Tarea
from .forms import VehiculoForm
from agenda.models import CitaMantenimiento


@login_required
def gestionar_vehiculos(request):
    """Panel de gestión para ver/editar/eliminar todos los vehículos (admins/jefes)."""
    if not (request.user.is_superuser or request.user.rol in ['admin', 'jefe_taller']):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('No tienes permisos para ver este panel.')

    vehiculos = Vehiculo.objects.all().order_by('-fecha_ingreso')
    return render(request, 'vehiculos/gestionar_vehiculos.html', {'vehiculos': vehiculos})


@login_required
def eliminar_vehiculo(request, vehiculo_id):
    """Eliminar un vehículo (POST)."""
    if not (request.user.is_superuser or request.user.rol in ['admin', 'jefe_taller']):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('No tienes permisos para eliminar vehículos.')

    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    if request.method == 'POST':
        vehiculo.delete()
        messages.success(request, f'El vehículo {vehiculo.patente} ha sido eliminado.')
        return redirect('gestionar_vehiculos')

    # Si no es POST, redirigir al detalle
    return redirect('detalle_vehiculo', vehiculo_id=vehiculo.id)


@login_required
def tablero_clinica(request):
    """Vista tipo tablero Kanban para gestión visual como clínica"""
    if not (request.user.is_superuser or request.user.rol in ['admin', 'jefe_taller', 'mecanico']):
        messages.error(request, 'No tienes acceso a esta vista.')
        return redirect('clinica_dashboard')
    
    # Vehículos en reparación: por estado o por tareas activas
    vehiculos_reparacion = Vehiculo.objects.filter(
        Q(estado='reparacion') |
        (
            Q(tareas__estado__in=['en_proceso', 'pendiente']) & ~Q(estado__in=['listo', 'entregado'])
        )
    ).select_related('mecanico_asignado').prefetch_related('tareas').distinct().order_by('-fecha_ingreso')

    # Vehículos listos para retiro
    vehiculos_listos = Vehiculo.objects.filter(estado='listo').select_related('mecanico_asignado').order_by('-fecha_ingreso')

    # (Opcional) mantener para compatibilidad aunque no se usen en la vista simplificada
    vehiculos_ingresados = Vehiculo.objects.filter(estado='ingresado').order_by('-fecha_ingreso')
    vehiculos_diagnostico = Vehiculo.objects.filter(estado='diagnostico').order_by('-fecha_ingreso')

    # Vistas simplificadas: Solicitudes (citas pendientes, aunque no tengan vehículo),
    # Asignación (confirmada sin mecánico), Listos
    citas_solicitudes = CitaMantenimiento.objects.filter(
        estado='pendiente'
    ).select_related('vehiculo', 'solicitante').order_by('fecha_hora')

    solicitudes = Vehiculo.objects.filter(
        citas__estado='pendiente'
    ).select_related('mecanico_asignado').distinct().order_by('-fecha_ingreso')
    
    # Estadísticas para cada columna
    stats = {
        'solicitudes': {
            'count': citas_solicitudes.count(),
            'color': '#ffc107',
            'icon': '🟡'
        },
        'listos': {
            'count': vehiculos_listos.count(),
            'color': '#007bff',
            'icon': '🔷'
        }
    }
    
    # Obtener mecánicos activos para jefes de taller
    mecanicos_activos = []
    if request.user.rol in ['jefe_taller', 'admin']:
        from usuarios.models import Usuario
        mecanicos_activos = Usuario.objects.filter(rol='mecanico', is_active=True)
    
    # Mapear cita activa (cualquier tipo) por vehículo para acciones de aprobación
    for v in list(vehiculos_reparacion) + list(vehiculos_listos):
        cita = CitaMantenimiento.objects.filter(
            vehiculo=v,
            estado__in=['pendiente', 'confirmada']
        ).order_by('fecha_hora').first()
        setattr(v, 'cita_activa', cita)

    context = {
        'ingresados': vehiculos_ingresados,
        'diagnostico': vehiculos_diagnostico,
        'reparacion': vehiculos_reparacion,
        'listos': vehiculos_listos,
        'citas_solicitudes': citas_solicitudes,
        'solicitudes': solicitudes,
        'stats': stats,
        'mecanicos_activos': mecanicos_activos,
    }
    return render(request, 'vehiculos/tablero_clinica.html', context)



@login_required
def cambiar_estado_vehiculo(request, vehiculo_id):
    """Cambiar estado del vehículo desde el tablero"""
    if request.method == 'POST':
        vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
        nuevo_estado = request.POST.get('nuevo_estado')
        mecanico_id = request.POST.get('mecanico_id')
        
        estado_anterior = vehiculo.estado
        
        # Validar permisos
        if request.user.rol not in ['jefe_taller', 'admin', 'mecanico']:
            messages.error(request, 'No tienes permisos para cambiar estados de vehículos.')
            return redirect('clinica_dashboard')
        
        if nuevo_estado in dict(Vehiculo.ESTADOS).keys():
            vehiculo.estado = nuevo_estado
            
            # Asignar mecánico si se proporciona
            if mecanico_id and request.user.rol in ['jefe_taller', 'admin']:
                from usuarios.models import Usuario
                try:
                    mecanico = Usuario.objects.get(id=mecanico_id, rol='mecanico')
                    vehiculo.mecanico_asignado = mecanico
                    messages.info(request, f'Vehículo asignado a {mecanico.get_full_name()}')
                except Usuario.DoesNotExist:
                    messages.error(request, 'Mecánico no válido.')
            
            # Auto-asignación si es mecánico y no hay mecánico asignado
            elif nuevo_estado == 'reparacion' and not vehiculo.mecanico_asignado and request.user.rol == 'mecanico':
                vehiculo.mecanico_asignado = request.user
                messages.info(request, f'Te has auto-asignado el vehículo {vehiculo.patente}')
            
            # Lógica adicional según el estado
            if nuevo_estado == 'listo' and not vehiculo.fecha_salida:
                vehiculo.fecha_salida = timezone.now()
                vehiculo.mecanico_asignado = None
                Tarea.objects.filter(vehiculo=vehiculo, estado__in=['en_proceso', 'pendiente', 'pausada']).update(
                    estado='completada',
                    fecha_fin=timezone.now()
                )
                messages.success(request, f'Vehículo {vehiculo.patente} marcado como listo para retiro!')
            
            elif nuevo_estado == 'entregado':
                messages.success(request, f'Vehículo {vehiculo.patente} marcado como entregado!')
            
            vehiculo.save()
            
            messages.success(
                request, 
                f'Vehículo {vehiculo.patente} cambiado de {vehiculo.get_estado_display()} a {vehiculo.get_estado_display()}'
            )
        else:
            messages.error(request, 'Estado inválido.')
    
    return redirect('clinica_dashboard')



@login_required
def lista_taller(request):
    vehiculos = Vehiculo.objects.exclude(estado='entregado').order_by('-fecha_ingreso')
    return render(request, 'vehiculos/lista_taller.html', {'vehiculos': vehiculos})

@login_required
def detalle_vehiculo(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    return render(request, 'vehiculos/detalle_vehiculo.html', {'vehiculo': vehiculo})# vehiculos/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from .models import Vehiculo
from .forms import VehiculoForm

@login_required
def clinica_dashboard(request):
    # Estadísticas en tiempo real
    total_vehiculos = Vehiculo.objects.exclude(estado__in=['pendiente', 'rechazado']).count()
    vehiculos_taller = Vehiculo.objects.exclude(estado__in=['entregado', 'pendiente', 'rechazado']).count()
    vehiculos_ingresados = Vehiculo.objects.filter(estado='ingresado').count()
    vehiculos_diagnostico = Vehiculo.objects.filter(estado='diagnostico').count()
    vehiculos_reparacion = Vehiculo.objects.filter(estado='reparacion').count()
    vehiculos_listos = Vehiculo.objects.filter(estado='listo').count()
    
    # Últimos vehículos ingresados
    ultimos_vehiculos = Vehiculo.objects.exclude(estado__in=['pendiente', 'rechazado']).order_by('-fecha_ingreso')[:5]
    
    context = {
        'total_vehiculos': total_vehiculos,
        'vehiculos_taller': vehiculos_taller,
        'vehiculos_ingresados': vehiculos_ingresados,
        'vehiculos_diagnostico': vehiculos_diagnostico,
        'vehiculos_reparacion': vehiculos_reparacion,
        'vehiculos_listos': vehiculos_listos,
        'ultimos_vehiculos': ultimos_vehiculos,
    }
    return render(request, 'vehiculos/clinica_dashboard.html', context)

@login_required
def ingreso_vehiculo(request):
    # Verificación de permisos: recepcionista, admin o chofer pueden registrar vehículos
    if request.user.rol not in ['recepcionista', 'admin', 'chofer']:
        messages.error(request, 'No tienes permiso para ingresar vehículos.')
        return redirect('clinica_dashboard')

    if request.method == 'POST':
        form = VehiculoForm(request.POST)
        if form.is_valid():
            vehiculo = form.save(commit=False)
            if request.user.rol == 'chofer':
                nombre = request.user.get_full_name() or request.user.username
                telefono = getattr(request.user, 'telefono', '') or ''
                empresa = 'PepsiCo Chile'
                vehiculo.nombre_chofer = nombre
                vehiculo.telefono_chofer = telefono
                vehiculo.empresa_chofer = empresa
                vehiculo.chofer_asignado = request.user
                vehiculo.estado = 'pendiente'
            else:
                vehiculo.guardia_ingreso = request.user  # Registrar quién ingresó el vehículo
            vehiculo.save()

            if request.user.rol == 'guardia':
                for archivo in request.FILES.getlist('fotos_ingreso'):
                    try:
                        doc = DocumentoVehiculo(
                            vehiculo=vehiculo,
                            tipo_documento='foto_ingreso',
                            archivo=archivo,
                            subido_por=request.user,
                            descripcion='Foto estado inicial'
                        )
                        doc.save()
                    except Exception as e:
                        messages.warning(request, f'Error al guardar una foto de ingreso: {str(e)}')
            elif request.FILES.getlist('fotos_ingreso'):
                messages.warning(request, 'Las fotos del estado inicial solo pueden ser registradas por el guardia.')

            # 🔗 Vincular citas pendientes
            citas_pendientes = CitaMantenimiento.objects.filter(
                patente=vehiculo.patente,
                estado='pendiente',
                vehiculo__isnull=True  # Solo actualizamos las que no tengan vehículo asignado
            )

            if citas_pendientes.exists():
                citas_pendientes.update(vehiculo=vehiculo, estado='confirmada')
                messages.success(request, f'Vehículo {vehiculo.patente} ingresado y cita vinculada.')
            else:
                messages.info(request, f'Vehículo {vehiculo.patente} ingresado, pero no se encontraron citas pendientes para vincular.')

            destino = 'clinica_dashboard'
            if request.user.is_superuser or request.user.rol == 'admin':
                destino = 'dashboard_admin'
            elif request.user.rol == 'jefe_taller':
                destino = 'dashboard_jefe_taller'
            elif request.user.rol == 'guardia':
                destino = 'dashboard_guardia'
            elif request.user.rol == 'recepcionista':
                destino = 'dashboard_recepcionista'
            elif request.user.rol == 'chofer':
                destino = 'dashboard_chofer'
            return redirect(destino)
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
            messages.error(request, 'Hubo un error al ingresar el vehículo. Verifica los datos.')

    else:
        if request.user.rol == 'chofer':
            nombre = request.user.get_full_name() or request.user.username
            telefono = getattr(request.user, 'telefono', '') or ''
            empresa = 'PepsiCo Chile'
            form = VehiculoForm(initial={
                'nombre_chofer': nombre,
                'telefono_chofer': telefono,
                'empresa_chofer': empresa,
            })
        else:
            form = VehiculoForm()

    return render(request, 'vehiculos/ingreso_vehiculo.html', {'form': form})


@login_required
def importar_vehiculos_excel(request):
    """Importar múltiples vehículos desde archivo Excel o CSV"""
    if request.user.rol not in ['admin', 'jefe_taller', 'guardia', 'recepcionista']:
        messages.error(request, 'No tienes permiso para importar vehículos.')
        return redirect('clinica_dashboard')
    
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        import csv
        import io
        from django.db import transaction
        
        archivo = request.FILES['archivo_excel']
        nombre_archivo = archivo.name.lower()
        
        try:
            vehiculos_creados = 0
            vehiculos_actualizados = 0
            errores = []
            
            # Determinar si es CSV o Excel
            if nombre_archivo.endswith('.csv'):
                # Procesar CSV
                archivo_texto = archivo.read().decode('utf-8-sig')
                csv_reader = csv.reader(io.StringIO(archivo_texto), delimiter=';')
                next(csv_reader)  # Saltar encabezado
                
                with transaction.atomic():
                    for idx, row in enumerate(csv_reader, start=2):
                        try:
                            if len(row) < 8:
                                continue
                                
                            patente, marca, modelo, año, tipo, flota, kilometraje, activo = row[:8]
                            
                            # Limpiar y validar datos
                            if not patente or not patente.strip():
                                continue
                                
                            patente = str(patente).strip().upper()
                            marca = str(marca).strip() if marca else ''
                            modelo = str(modelo).strip() if modelo else ''
                            
                            try:
                                año = int(año) if año and año.strip() else None
                            except:
                                año = None
                            
                            flota = str(flota).strip() if flota and str(flota).strip() not in ['#N/D', 'N/D', ''] else ''
                            
                            # Convertir tipo a formato del modelo
                            tipo_map = {
                                'Funcional Flota': 'funcional_flota',
                                'Camion DTS': 'camion_dts',
                                'Camion Cstore': 'camion_cstore',
                                'MERCHANDISING': 'merchandising',
                                'Camión': 'camion',
                                'Furgón': 'furgon',
                            }
                            tipo_vehiculo = tipo_map.get(tipo, 'automovil')
                            
                            # Convertir kilometraje
                            if kilometraje and str(kilometraje).strip():
                                try:
                                    kilometraje = int(kilometraje)
                                except:
                                    kilometraje = 0
                            else:
                                kilometraje = 0
                            
                            # Convertir activo
                            activo_bool = str(activo).upper() in ['SI', 'SÍ', 'YES', 'TRUE', '1']
                            
                            # Crear o actualizar vehículo
                            vehiculo, created = Vehiculo.objects.update_or_create(
                                patente=patente,
                                defaults={
                                    'marca': marca,
                                    'modelo': modelo,
                                    'año': año,
                                    'tipo_vehiculo': tipo_vehiculo,
                                    'flota': flota,
                                    'kilometraje': kilometraje,
                                    'activo': activo_bool,
                                    'motivo_ingreso': 'Importado desde archivo',
                                    'guardia_ingreso': request.user,
                                }
                            )
                            
                            if created:
                                vehiculos_creados += 1
                            else:
                                vehiculos_actualizados += 1
                                
                        except Exception as e:
                            errores.append(f"Fila {idx}: {str(e)}")
            
            elif nombre_archivo.endswith(('.xlsx', '.xls')):
                # Procesar Excel
                import openpyxl
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active
                
                with transaction.atomic():
                    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            if not row or len(row) < 8:
                                continue
                                
                            patente, marca, modelo, año, tipo, flota, kilometraje, activo = row[:8]
                            
                            # Limpiar y validar datos
                            if not patente:
                                continue
                                
                            patente = str(patente).strip().upper()
                            marca = str(marca).strip() if marca else ''
                            modelo = str(modelo).strip() if modelo else ''
                            año = int(año) if año else None
                            flota = str(flota).strip() if flota and str(flota).strip() not in ['#N/D', 'N/D'] else ''
                            
                            # Convertir tipo a formato del modelo
                            tipo_map = {
                                'Funcional Flota': 'funcional_flota',
                                'Camion DTS': 'camion_dts',
                                'Camion Cstore': 'camion_cstore',
                                'MERCHANDISING': 'merchandising',
                                'Camión': 'camion',
                                'Furgón': 'furgon',
                            }
                            tipo_vehiculo = tipo_map.get(tipo, 'automovil')
                            
                            # Convertir kilometraje
                            if kilometraje and str(kilometraje).strip():
                                try:
                                    kilometraje = int(kilometraje)
                                except:
                                    kilometraje = 0
                            else:
                                kilometraje = 0
                            
                            # Convertir activo
                            activo_bool = str(activo).upper() in ['SI', 'SÍ', 'YES', 'TRUE', '1']
                            
                            # Crear o actualizar vehículo
                            vehiculo, created = Vehiculo.objects.update_or_create(
                                patente=patente,
                                defaults={
                                    'marca': marca,
                                    'modelo': modelo,
                                    'año': año,
                                    'tipo_vehiculo': tipo_vehiculo,
                                    'flota': flota,
                                    'kilometraje': kilometraje,
                                    'activo': activo_bool,
                                    'motivo_ingreso': 'Importado desde Excel',
                                    'guardia_ingreso': request.user,
                                }
                            )
                            
                            if created:
                                vehiculos_creados += 1
                            else:
                                vehiculos_actualizados += 1
                                
                        except Exception as e:
                            errores.append(f"Fila {idx}: {str(e)}")
            else:
                messages.error(request, 'Formato de archivo no soportado. Use .xlsx, .xls o .csv')
                return redirect('importar_vehiculos_excel')
            
            # Mensajes de resultado
            if vehiculos_creados > 0:
                messages.success(request, f'✅ {vehiculos_creados} vehículos creados exitosamente')
            if vehiculos_actualizados > 0:
                messages.info(request, f'📝 {vehiculos_actualizados} vehículos actualizados')
            if errores:
                for error in errores[:5]:  # Mostrar solo los primeros 5 errores
                    messages.warning(request, error)
                if len(errores) > 5:
                    messages.warning(request, f'... y {len(errores) - 5} errores más')
            
            return redirect('gestionar_vehiculos')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('importar_vehiculos_excel')
    
    return render(request, 'vehiculos/importar_excel.html')


@login_required
def exportar_vehiculos_excel(request):
    """Exporta todos los vehículos a un archivo Excel descargable (.xlsx)."""
    if not (request.user.is_superuser or request.user.rol in ['admin', 'jefe_taller']):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('No tienes permisos para exportar vehículos.')

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        vehiculos = Vehiculo.objects.all().order_by('id')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Vehículos'

        headers = [
            'ID', 'Patente', 'Marca', 'Modelo', 'Año', 'Tipo Vehículo', 'Flota',
            'Kilometraje', 'Activo', 'Nombre Chofer', 'Teléfono Chofer', 'Empresa Chofer',
            'Estado', 'Fecha Ingreso'
        ]
        ws.append(headers)

        for v in vehiculos:
            tipo_display = getattr(v, 'get_tipo_vehiculo_display', None)
            tipo_val = tipo_display() if callable(tipo_display) else (v.tipo_vehiculo if hasattr(v, 'tipo_vehiculo') else '')
            row = [
                v.id,
                v.patente,
                v.marca or '',
                v.modelo or '',
                v.año or '',
                tipo_val,
                v.flota or '',
                v.kilometraje or 0,
                'SI' if getattr(v, 'activo', False) else 'NO',
                v.nombre_chofer or '',
                v.telefono_chofer or '',
                v.empresa_chofer or '',
                v.get_estado_display() if hasattr(v, 'get_estado_display') else (v.estado if hasattr(v, 'estado') else ''),
                v.fecha_ingreso.strftime('%d/%m/%Y %H:%M') if getattr(v, 'fecha_ingreso', None) else ''
            ]
            ws.append(row)

        # Ajustar ancho de columnas
        for i, col in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ''
                except Exception:
                    val = ''
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"vehiculos_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar el Excel: {str(e)}')
        return redirect('gestionar_vehiculos')

@login_required
def lista_taller(request):
    vehiculos = Vehiculo.objects.exclude(estado__in=['entregado', 'pendiente', 'rechazado']).order_by('-fecha_ingreso')
    return render(request, 'vehiculos/lista_taller.html', {'vehiculos': vehiculos})

@login_required
def detalle_vehiculo(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    return render(request, 'vehiculos/detalle_vehiculo.html', {'vehiculo': vehiculo})

@login_required
def confirmar_ingreso_vehiculo(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)

    if request.method != 'POST':
        messages.error(request, 'Método no válido para confirmar ingreso.')
        return redirect('detalle_vehiculo', vehiculo_id=vehiculo.id)

    if not (request.user.is_superuser or request.user.rol in ['admin', 'guardia', 'recepcionista']):
        messages.error(request, 'No tienes permisos para confirmar el ingreso.')
        return redirect('detalle_vehiculo', vehiculo_id=vehiculo.id)

    estado_anterior = vehiculo.estado

    from django.utils import timezone
    vehiculo.estado = 'ingresado'
    vehiculo.fecha_ingreso = timezone.now()
    if request.user.rol in ['guardia', 'recepcionista']:
        vehiculo.guardia_ingreso = request.user
    vehiculo.save()

    try:
        citas_pendientes = CitaMantenimiento.objects.filter(
            patente=vehiculo.patente,
            estado='pendiente',
            vehiculo__isnull=True
        )
        if citas_pendientes.exists():
            citas_pendientes.update(vehiculo=vehiculo, estado='confirmada')
            messages.info(request, 'Cita vinculada y confirmada para este vehículo.')
        else:
            messages.warning(request, 'No se encontró cita pendiente. Verifica autorización de ingreso.')
    except Exception as e:
        messages.warning(request, f'No fue posible verificar citas: {str(e)}')

    messages.success(request, f'Ingreso confirmado. Estado {estado_anterior} → Ingresado.')
    return redirect('detalle_vehiculo', vehiculo_id=vehiculo.id)

@login_required
def ignorar_ingreso_vehiculo(request, vehiculo_id):
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)

    if request.method != 'POST':
        messages.error(request, 'Método no válido para ignorar ingreso.')
        return redirect('detalle_vehiculo', vehiculo_id=vehiculo.id)

    if not (request.user.is_superuser or request.user.rol in ['admin', 'guardia', 'recepcionista']):
        messages.error(request, 'No tienes permisos para ignorar ingresos.')
        return redirect('detalle_vehiculo', vehiculo_id=vehiculo.id)

    vehiculo.estado = 'rechazado'
    vehiculo.save()
    messages.info(request, 'Solicitud de ingreso ignorada. El vehículo no aparecerá en pendientes.')
    return redirect('pendientes_ingreso')

@login_required
def busqueda_avanzada(request):
    query = request.GET.get('q', '').strip()
    resultados = {
        'vehiculos': [],
        'repuestos': [], 
        'tareas': [],
    }
    
    if query:
        # Búsqueda en vehículos
        resultados['vehiculos'] = Vehiculo.objects.filter(
            Q(patente__icontains=query) |
            Q(marca__icontains=query) |
            Q(modelo__icontains=query) |
            Q(nombre_chofer__icontains=query) |
            Q(numero_chasis__icontains=query)
        ).select_related('mecanico_asignado').order_by('-fecha_ingreso')[:10]
        
        # Búsqueda en repuestos
        resultados['repuestos'] = Repuesto.objects.filter(
            Q(codigo__icontains=query) |
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(modelo_compatible__icontains=query) |
            Q(proveedor__icontains=query)
        ).select_related('categoria').order_by('nombre')[:10]
        
        # Búsqueda en tareas
        resultados['tareas'] = Tarea.objects.filter(
            Q(titulo__icontains=query) |
            Q(descripcion__icontains=query)
        ).select_related('vehiculo', 'mecanico_asignado').order_by('-fecha_creacion')[:10]
    
    context = {
        'query': query,
        'resultados': resultados,
        'total_resultados': sum(len(v) for v in resultados.values())
    }
    return render(request, 'vehiculos/busqueda_avanzada.html', context)


@login_required
def pendientes_ingreso(request):
    if not (request.user.is_superuser or request.user.rol in ['admin', 'guardia', 'recepcionista']):
        messages.error(request, 'No tienes permisos para ver pendientes de ingreso.')
        return redirect('clinica_dashboard')

    vehiculos = Vehiculo.objects.filter(estado='pendiente').annotate(
        fotos_ingreso=Count('documentos', filter=Q(documentos__tipo_documento='foto_ingreso'))
    ).order_by('-fecha_ingreso')

    return render(request, 'vehiculos/pendientes_ingreso.html', {
        'vehiculos': vehiculos
    })

@login_required
def editar_vehiculo(request, vehiculo_id):
    """Editar un vehículo. Solo accesible para jefes de taller y admins."""
    if request.user.rol not in ['jefe_taller', 'admin']:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('No tienes permisos para editar vehículos.')

    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    if request.method == 'POST':
        form = VehiculoForm(request.POST, instance=vehiculo)
        if form.is_valid():
            form.save()
            from django.contrib import messages
            messages.success(request, f'Vehículo {vehiculo.patente} actualizado correctamente.')
            return redirect('detalle_vehiculo', vehiculo_id=vehiculo.id)
    else:
        form = VehiculoForm(instance=vehiculo)

    return render(request, 'vehiculos/ingreso_vehiculo.html', {'form': form, 'editar': True, 'vehiculo': vehiculo})

@login_required
def vehiculo_por_patente_json(request):
    patente = (request.GET.get('patente') or '').strip()
    if not patente:
        return JsonResponse({'found': False, 'error': 'patente requerida'}, status=400)
    v = Vehiculo.objects.filter(patente__iexact=patente).select_related('mecanico_asignado').first()
    if not v:
        return JsonResponse({'found': False})
    data = {
        'found': True,
        'id': v.id,
        'patente': v.patente,
        'marca': v.marca or '',
        'modelo': v.modelo or '',
        'estado': getattr(v, 'get_estado_display')() if hasattr(v, 'get_estado_display') else (v.estado or ''),
        'mecanico': (v.mecanico_asignado.get_full_name() or v.mecanico_asignado.username) if v.mecanico_asignado else '',
        'fecha_ingreso': v.fecha_ingreso.strftime('%d/%m/%Y %H:%M') if getattr(v, 'fecha_ingreso', None) else '',
    }
    return JsonResponse(data)
