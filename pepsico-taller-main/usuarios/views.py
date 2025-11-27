# usuarios/views.py
from django.shortcuts import render, redirect, get_object_or_404 
from django.contrib.auth import login as auth_login, authenticate, logout as auth_logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.http import HttpResponse
from .forms import LoginForm, RegistroUsuarioForm, PerfilUsuarioForm

# Importar modelos de otras apps
from django.db import models
from vehiculos.models import Vehiculo, DocumentoVehiculo, Siniestro
from mantenimientos.models import Tarea, Pausa  # ← Agrega "Pausa" aquí
from inventario.models import Repuesto, MovimientoInventario
from agenda.models import CitaMantenimiento 
from usuarios.models import Usuario  # ← Importar Usuario desde la misma app
from django.views.decorators.http import require_http_methods

@login_required
@require_http_methods(["GET", "POST"])
def role_selector(request):
    """Permite al usuario con suficientes permisos ver la aplicación como otro rol.

    - GET: muestra un formulario con los roles disponibles.
    - POST: actualiza la sesión con el rol seleccionado y redirige al dashboard.
    """
    # Solo administradores y jefes de taller pueden activar la vista como rol
    if not (request.user.is_superuser or request.user.rol == 'admin' or request.user.rol == 'jefe_taller'):
        messages.error(request, 'No tienes permisos para usar "Vista como Rol".')
        return redirect('clinica_dashboard')

    roles = Usuario.ROLES

    if request.method == 'POST':
        view_as = request.POST.get('rol')
        valid_roles = [r for r, _ in roles]
        if view_as not in valid_roles:
            messages.error(request, 'Rol inválido.')
            return redirect('role_selector')

        request.session['view_as_role'] = view_as
        # Guardar usuario original para mostrar en la interfaz
        request.session['original_user_id'] = request.user.id
        messages.success(request, f'Ahora estás viendo la aplicación como: {dict(roles).get(view_as)}')
        
        # Crear un usuario temporal con el rol seleccionado para redirigir al dashboard correcto
        from copy import copy
        temp_user = copy(request.user)
        temp_user.rol = view_as
        return redirect_a_dashboard_especifico(temp_user)

    return render(request, 'usuarios/role_selector.html', {'roles': roles})


@login_required
def clear_view_as_role(request):
    """Limpiar la sesión de 'view as role' para regresar al usuario real."""
    if 'view_as_role' in request.session:
        del request.session['view_as_role']
    if 'original_user_id' in request.session:
        del request.session['original_user_id']
    
    # Forzar guardado de la sesión
    request.session.modified = True
    
    messages.success(request, 'Has salido de la vista por rol. Volviendo a tu rol original.')
    
    # Redirigir al dashboard del rol REAL del usuario
    return redirect_a_dashboard_especifico(request.user)

def custom_login(request):
    """Vista personalizada de login con redirección a dashboard específico"""
    if request.user.is_authenticated:
        print(f"✅ Usuario YA autenticado: {request.user.username}, Rol: {request.user.rol}")  # ← DEPURACIÓN
        return redirect_a_dashboard_especifico(request.user)
    
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                auth_login(request, user)
                print(f"✅ Login EXITOSO: {user.username}, Rol: {user.rol}")  # ← DEPURACIÓN
                messages.success(request, f'¡Bienvenido {user.get_full_name() or user.username}!')
                return redirect_a_dashboard_especifico(user)
            else:
                messages.error(request, 'Usuario o contraseña incorrectos.')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = LoginForm()
    
    return render(request, 'usuarios/login.html', {'form': form})


#función redirect_a_dashboard_especifico
def redirect_a_dashboard_especifico(user):
    """Redirigir al usuario a su dashboard específico según su rol"""
    print(f"🔍 Redirigiendo usuario: {user.username}, Rol: {user.rol}")  # ← DEPURACIÓN

    if user.rol == 'admin':
        print("📍 Redirigiendo a dashboard_admin")  # ← DEPURACIÓN OPCIONAL
        return redirect('dashboard_admin') # <-- AÑADIR ESTA LÍNEA
    elif user.rol == 'mecanico':
        print("📍 Redirigiendo a dashboard_mecanico")  # ← DEPURACIÓN
        return redirect('dashboard_mecanico')
    elif user.rol == 'jefe_taller':
        return redirect('dashboard_jefe_taller')
    elif user.rol == 'chofer':
        print("📍 Redirigiendo a dashboard_chofer")  # ← DEPURACIÓN
        return redirect('dashboard_chofer')
    elif user.rol == 'vendedor':
        print("📍 Redirigiendo a dashboard_vendedor")  # ← DEPURACIÓN
        return redirect('dashboard_vendedor')
    elif user.rol == 'bodeguero':
        return redirect('dashboard_bodeguero')
    elif user.rol == 'guardia':
        print("📍 Redirigiendo a dashboard_guardia")  # ← DEPURACIÓN
        return redirect('dashboard_guardia')
    elif user.rol == 'ehs':
        print("📍 Redirigiendo a dashboard_ehs")  # ← DEPURACIÓN
        return redirect('dashboard_ehs')
    elif user.rol == 'recepcionista':
        return redirect('dashboard_recepcionista')
    else:
        print("📍 Redirigiendo a clinica_dashboard (rol no específico o no mapeado)") 
        return redirect('clinica_dashboard') # <-- 'admin' ya no entra aquí


@login_required
def registro_usuario(request):
    """Registro de nuevos usuarios - Solo para administradores"""
    # Verificar que el usuario sea administrador
    if not (request.user.is_superuser or request.user.rol == 'admin'):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('clinica_dashboard')
    
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            messages.success(request, f'¡Usuario {user.username} creado exitosamente! Correo: {user.email}')
            # Redirigir al dashboard de admin en lugar de iniciar sesión como el nuevo usuario
            return redirect('dashboard_admin')
    else:
        form = RegistroUsuarioForm()
    
    return render(request, 'usuarios/registro.html', {'form': form})

@login_required
def lista_usuarios(request):
    """Lista de todos los usuarios - Solo para administradores"""
    # Verificar que el usuario sea administrador
    if not (request.user.is_superuser or request.user.rol == 'admin'):
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('clinica_dashboard')
    
    usuarios = Usuario.objects.all().order_by('-date_joined')
    admin_count = Usuario.objects.filter(models.Q(rol='admin') | models.Q(is_superuser=True)).count()
    mecanicos_count = Usuario.objects.filter(rol='mecanico').count()
    active_count = Usuario.objects.filter(is_active=True).count()
    context = {
        'usuarios': usuarios,
        'admin_count': admin_count,
        'mecanicos_count': mecanicos_count,
        'active_count': active_count,
    }
    return render(request, 'usuarios/lista_usuarios.html', context)

@login_required
def exportar_usuarios_excel(request):
    """Exportar lista de usuarios a Excel - Solo para administradores"""
    # Verificar que el usuario sea administrador
    if not (request.user.is_superuser or request.user.rol == 'admin'):
        messages.error(request, 'No tienes permisos para exportar esta información.')
        return redirect('lista_usuarios')
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
    except ImportError as e:
        messages.error(request, f'Error al importar librerías necesarias: {str(e)}')
        return redirect('lista_usuarios')
    
    try:
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios Sistema PepsiCo"
        
        # Colores corporativos PepsiCo
        PEPSICO_BLUE = "004B93"
        PEPSICO_RED = "E32934"
        PEPSICO_LIGHT_BLUE = "5B9BD5"
        LIGHT_GRAY = "F2F2F2"
        WHITE = "FFFFFF"
        
        # Estilos mejorados
        # Título principal
        title_font = Font(bold=True, size=18, color=PEPSICO_BLUE)
        title_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        
        # Headers
        header_font = Font(bold=True, size=12, color=WHITE)
        header_fill = PatternFill(start_color=PEPSICO_BLUE, end_color=PEPSICO_BLUE, fill_type="solid")
        
        # Filas alternadas
        row_fill_1 = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        row_fill_2 = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        
        # Estados específicos
        active_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        inactive_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        admin_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )
        
        thick_border = Border(
            left=Side(style='medium', color=PEPSICO_BLUE),
            right=Side(style='medium', color=PEPSICO_BLUE),
            top=Side(style='medium', color=PEPSICO_BLUE),
            bottom=Side(style='medium', color=PEPSICO_BLUE)
        )
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Agregar título principal
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = f"📊 REPORTE DE USUARIOS - SISTEMA PEPSICO TALLER"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_alignment
        title_cell.border = thick_border
        
        # Agregar información de generación
        ws.merge_cells('A2:L2')
        info_cell = ws['A2']
        info_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')} | Total de usuarios: {Usuario.objects.count()}"
        info_cell.font = Font(size=10, italic=True, color="666666")
        info_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        info_cell.alignment = center_alignment
        info_cell.border = thin_border
        
        # Fila vacía para separación
        ws.row_dimensions[3].height = 5
        
        # Headers (ahora en la fila 4)
        headers = [
            'ID', 'Usuario', 'Nombre', 'Apellido', 'Nombre Completo',
            'Correo Electrónico', 'Teléfono', 'Rol', 'Estado',
            'Es Superusuario', 'Fecha Registro', 'Último Acceso'
        ]
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thick_border
            
        # Aumentar altura de la fila de headers
        ws.row_dimensions[4].height = 25
        
        # Obtener datos de usuarios
        usuarios = Usuario.objects.all().order_by('-date_joined')
        
        # Escribir datos (comenzando desde la fila 5)
        for idx, usuario in enumerate(usuarios):
            row = idx + 5  # Comenzar desde la fila 5
            
            # Determinar color de fila alternada
            row_fill = row_fill_1 if idx % 2 == 0 else row_fill_2
            
            data = [
                usuario.id,
                usuario.username,
                usuario.first_name,
                usuario.last_name,
                usuario.get_full_name() or 'Sin nombre',
                usuario.email,
                usuario.telefono or 'No especificado',
                usuario.get_rol_display(),
                '✅ Activo' if usuario.is_active else '❌ Inactivo',
                '👑 Sí' if usuario.is_superuser else 'No',
                usuario.date_joined.strftime('%d/%m/%Y %H:%M') if usuario.date_joined else 'N/A',
                usuario.last_login.strftime('%d/%m/%Y %H:%M') if usuario.last_login else 'Nunca se conectó'
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                # Alineación específica por columna
                if col in [1]:  # ID
                    cell.alignment = center_alignment
                elif col in [5, 6]:  # Nombre completo, Email
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment
                
                # Aplicar color de fondo especial según el contenido
                if col == 9:  # Columna Estado
                    if usuario.is_active:
                        cell.fill = active_fill
                        cell.font = Font(bold=True, color="2E7D32")
                    else:
                        cell.fill = inactive_fill
                        cell.font = Font(bold=True, color="C62828")
                elif col == 10 and usuario.is_superuser:  # Columna Superusuario
                    cell.fill = admin_fill
                    cell.font = Font(bold=True, color="E65100")
                elif col == 8:  # Columna Rol
                    # Colores específicos por rol
                    if usuario.rol == 'admin':
                        cell.font = Font(bold=True, color="C62828")
                    elif usuario.rol == 'jefe_taller':
                        cell.font = Font(bold=True, color="F57C00")
                    elif usuario.rol == 'mecanico':
                        cell.font = Font(bold=True, color="1565C0")
                    else:
                        cell.font = Font(color="424242")
                    cell.fill = row_fill
                else:
                    cell.fill = row_fill
                    cell.font = Font(color="424242")
            
            # Ajustar altura de cada fila de datos
            ws.row_dimensions[row].height = 20
        
        # Agregar pie de página con estadísticas
        last_row = len(usuarios) + 6
        
        # Fila vacía de separación
        ws.row_dimensions[last_row].height = 5
        last_row += 1
        
        # Estadísticas del reporte
        ws.merge_cells(f'A{last_row}:L{last_row}')
        stats_cell = ws[f'A{last_row}']
        
        total_usuarios = usuarios.count()
        usuarios_activos = usuarios.filter(is_active=True).count()
        usuarios_inactivos = usuarios.filter(is_active=False).count()
        superusuarios = usuarios.filter(is_superuser=True).count()
        
        stats_text = f"📈 ESTADÍSTICAS: Total: {total_usuarios} | Activos: {usuarios_activos} | Inactivos: {usuarios_inactivos} | Superusuarios: {superusuarios}"
        
        stats_cell.value = stats_text
        stats_cell.font = Font(bold=True, size=11, color=PEPSICO_BLUE)
        stats_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        stats_cell.alignment = center_alignment
        stats_cell.border = thick_border
        
        # Ajustar ancho de columnas mejorado
        column_widths = [6, 18, 18, 18, 25, 35, 18, 20, 15, 18, 22, 22]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Congelar paneles (título y headers siempre visibles)
        ws.freeze_panes = ws['A5']
        
        # Agregar filtros automáticos a los headers
        ws.auto_filter.ref = f"A4:L{len(usuarios) + 4}"
        
        # Configuración de página para impresión
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # Márgenes
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
    
        # Configurar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'usuarios_pepsico_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar workbook en respuesta
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al generar el archivo Excel: {str(e)}')
        return redirect('lista_usuarios')

@login_required
def perfil_usuario(request):
    """
    Vista para editar el perfil del usuario y mostrar estadísticas específicas según su rol.
    Si el rol cambia, redirige automáticamente al nuevo dashboard.
    """
    user = request.user
    rol_original = user.rol  # Guardamos el rol ANTES de cualquier cambio

    # === Cargar datos específicos por rol (solo para mostrar en el perfil) ===
    datos_especificos = {}
    hoy = timezone.now().date()
    
    try:
        if user.rol == 'mecanico':
            tareas_asignadas = Tarea.objects.filter(mecanico_asignado=user)
            tareas_completadas = tareas_asignadas.filter(estado='completada')
            tareas_en_proceso = tareas_asignadas.filter(estado='en_proceso')
            datos_especificos = {
                'tipo': 'mecanico',
                'tareas_totales': tareas_asignadas.count(),
                'tareas_completadas': tareas_completadas.count(),
                'tareas_en_proceso': tareas_en_proceso.count(),
                'eficiencia': round((tareas_completadas.count() / tareas_asignadas.count() * 100), 1) if tareas_asignadas.count() > 0 else 0,
            }

        elif user.rol == 'chofer':
            # Buscar por relación directa o por nombre (compatibilidad)
            vehiculos_chofer = Vehiculo.objects.filter(chofer_asignado=user)
            if not vehiculos_chofer.exists():
                nombre_chofer = user.get_full_name() or user.username
                vehiculos_chofer = Vehiculo.objects.filter(
                    nombre_chofer__icontains=nombre_chofer
                )
            datos_especificos = {
                'tipo': 'chofer',
                'vehiculos_registrados': vehiculos_chofer.count(),
                'ultimo_ingreso': vehiculos_chofer.order_by('-fecha_ingreso').first(),
            }

        elif user.rol == 'guardia':
            vehiculos_ingresados = Vehiculo.objects.filter(guardia_ingreso=user)
            datos_especificos = {
                'tipo': 'guardia',
                'vehiculos_ingresados': vehiculos_ingresados.count(),
                'ingresos_hoy': vehiculos_ingresados.filter(fecha_ingreso__date=hoy).count(),
            }

        elif user.rol == 'jefe_taller':
            vehiculos_taller = Vehiculo.objects.exclude(estado='entregado')
            tareas_pendientes = Tarea.objects.filter(estado__in=['pendiente', 'en_proceso'])
            mecanicos_activos = Usuario.objects.filter(rol='mecanico', is_active=True).count()
            datos_especificos = {
                'tipo': 'supervisor',
                'vehiculos_taller': vehiculos_taller.count(),
                'tareas_pendientes': tareas_pendientes.count(),
                'mecanicos_activos': mecanicos_activos,
            }

        elif user.rol == 'ehs':
            tareas_riesgo = Tarea.objects.filter(prioridad='alta')
            siniestros_activos = Siniestro.objects.filter(estado__in=['reportado', 'en_evaluacion']).count()
            datos_especificos = {
                'tipo': 'ehs',
                'tareas_alto_riesgo': tareas_riesgo.count(),
                'siniestros_activos': siniestros_activos,
            }

        elif user.rol == 'bodeguero':
            from inventario.models import Repuesto, MovimientoInventario
            repuestos_bajo_stock = Repuesto.objects.filter(estado='bajo_stock').count()
            repuestos_agotados = Repuesto.objects.filter(estado='agotado').count()
            movimientos_hoy = MovimientoInventario.objects.filter(
                fecha_movimiento__date=hoy
            ).count()
            datos_especificos = {
                'tipo': 'bodeguero',
                'repuestos_bajo_stock': repuestos_bajo_stock,
                'repuestos_agotados': repuestos_agotados,
                'movimientos_hoy': movimientos_hoy,
            }

        elif user.rol == 'vendedor':
            vehiculos_asignados = Vehiculo.objects.filter(chofer_asignado=user)
            if not vehiculos_asignados.exists():
                nombre_vendedor = user.get_full_name() or user.username
                vehiculos_asignados = Vehiculo.objects.filter(
                    nombre_chofer__icontains=nombre_vendedor
                )

            citas_proximas = CitaMantenimiento.objects.filter(
                solicitante=user,
                fecha_hora__date__range=[hoy, hoy + timezone.timedelta(days=7)],
                estado__in=['pendiente', 'confirmada']
            )

            documentos_proximos = DocumentoVehiculo.objects.filter(
                vehiculo__in=vehiculos_asignados,
                fecha_vencimiento__gte=hoy,
                fecha_vencimiento__lte=hoy + timezone.timedelta(days=30)
            )

            datos_especificos = {
                'tipo': 'vendedor',
                'vehiculos_asignados': vehiculos_asignados.count(),
                'citas_proximas': citas_proximas.count(),
                'documentos_proximos': documentos_proximos.count(),
                'ultima_cita': citas_proximas.first(),
            }

        # Roles sin datos específicos (admin, bodeguero, etc.)
        else:
            datos_especificos = {'tipo': 'basico'}

    except Exception as e:
        datos_especificos = {'tipo': 'basico', 'error': str(e)}

    # === Manejo del formulario de edición ===
    if request.method == 'POST':
        form = PerfilUsuarioForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            # Recargar el usuario desde la BD para obtener el rol actualizado
            user.refresh_from_db()
            messages.success(request, '¡Perfil actualizado exitosamente!')

            # ✅ Si el rol cambió, redirigir al nuevo dashboard
            if user.rol != rol_original:
                messages.info(
                    request,
                    f'Se ha cambiado tu rol a "{user.get_rol_display()}" y se ha redirigido a tu nuevo dashboard.'
                )
                return redirect_a_dashboard_especifico(user)
            else:
                return redirect('perfil_usuario')
    else:
        form = PerfilUsuarioForm(instance=user)

    return render(request, 'usuarios/perfil.html', {
        'form': form,
        'datos_especificos': datos_especificos
    })

def custom_logout(request):
    """Cerrar sesión personalizado"""
    auth_logout(request)
    messages.success(request, '¡Has cerrado sesión exitosamente!')
    return redirect('login')


# DASHBOARDS ESPECÍFICOS
@login_required
def dashboard_mecanico(request):
    """Dashboard específico para mecánicos - Solo ven sus tareas"""
    if not (request.user.rol == 'mecanico' or request.user.rol == 'admin' or request.user.is_superuser):
        messages.error(request, 'No tienes acceso a esta vista.')
        return redirect('clinica_dashboard')
    
    # Tareas del mecánico actual
    tareas = Tarea.objects.filter(mecanico_asignado=request.user).select_related(
        'vehiculo'
    ).order_by('-fecha_creacion')
    
    # Filtrar por estado si se especifica
    estado = request.GET.get('estado')
    if estado:
        tareas = tareas.filter(estado=estado)
    
    # Estadísticas rápidas
    estadisticas = {
        'total_tareas': tareas.count(),
        'tareas_pendientes': tareas.filter(estado='pendiente').count(),
        'tareas_proceso': tareas.filter(estado='en_proceso').count(),
        'tareas_pausadas': tareas.filter(estado='pausada').count(),
        'tareas_completadas': tareas.filter(estado='completada').count(),
    }
    
    # Tareas urgentes (alta prioridad)
    tareas_urgentes = tareas.filter(prioridad='alta', estado__in=['pendiente', 'en_proceso'])
    
    # Pausas activas
    pausas_activas = Pausa.objects.filter(
        tarea__mecanico_asignado=request.user, 
        fecha_fin__isnull=True
    ).select_related('tarea')
    
    context = {
        'tareas': tareas,
        'estadisticas': estadisticas,
        'tareas_urgentes': tareas_urgentes,
        'pausas_activas': pausas_activas,
        'filtro_estado': estado,
    }
    
    return render(request, 'usuarios/dashboard_mecanico.html', context)

# usuarios/views.py - MODIFICAR dashboard_chofer
@login_required
def dashboard_chofer(request):
    """Dashboard específico para choferes - Solo ven sus vehículos"""
    if not (request.user.rol == 'chofer' or request.user.rol == 'admin' or request.user.is_superuser):
        messages.error(request, 'No tienes acceso a esta vista.')
        return redirect('clinica_dashboard')
    
    # Buscar vehículos del chofer por chofer_asignado (NUEVO)
    vehiculos = Vehiculo.objects.filter(
        chofer_asignado=request.user
    ).order_by('-fecha_ingreso')
    
    # También buscar por nombre por compatibilidad
    if not vehiculos.exists():
        nombre_chofer = request.user.get_full_name() or request.user.username
        vehiculos = Vehiculo.objects.filter(
            nombre_chofer__icontains=nombre_chofer
        ).order_by('-fecha_ingreso')
    
    # Estadísticas del chofer
    estadisticas = {
        'total_vehiculos': vehiculos.count(),
        'en_taller': vehiculos.exclude(estado='entregado').count(),
        'listos_retiro': vehiculos.filter(estado='listo').count(),
        'entregados': vehiculos.filter(estado='entregado').count(),
    }
    
    # Último vehículo ingresado
    ultimo_ingreso = vehiculos.first()
    
    # Vehículos actualmente en taller
    vehiculos_taller = vehiculos.exclude(estado='entregado')
    
    context = {
        'vehiculos': vehiculos,
        'vehiculos_taller': vehiculos_taller,
        'estadisticas': estadisticas,
        'ultimo_ingreso': ultimo_ingreso,
        'nombre_chofer': request.user.get_full_name() or request.user.username,
    }
    
    return render(request, 'usuarios/dashboard_chofer.html', context)


@login_required
def dashboard_guardia(request):
    """Dashboard específico para guardias - Solo ingreso de vehículos"""
    if not (request.user.rol == 'guardia' or request.user.rol == 'admin' or request.user.is_superuser):
        messages.error(request, 'No tienes acceso a esta vista.')
        return redirect('clinica_dashboard')
    
    # Vehículos ingresados por este guardia
    vehiculos_ingresados = Vehiculo.objects.filter(
        guardia_ingreso=request.user
    ).order_by('-fecha_ingreso')
    
    # Estadísticas
    hoy = timezone.now().date()
    estadisticas = {
        'total_ingresados': vehiculos_ingresados.count(),
        'ingresos_hoy': vehiculos_ingresados.filter(fecha_ingreso__date=hoy).count(),
        'ingresos_semana': vehiculos_ingresados.filter(
            fecha_ingreso__date__gte=hoy - timezone.timedelta(days=7)
        ).count(),
        'en_taller_actual': Vehiculo.objects.exclude(estado__in=['entregado', 'pendiente', 'rechazado']).count(),
    }
    
    # Últimos ingresos
    ultimos_ingresos = vehiculos_ingresados[:10]
    
    # Vehículos en taller actualmente
    vehiculos_taller = Vehiculo.objects.exclude(estado__in=['entregado', 'pendiente', 'rechazado']).order_by('-fecha_ingreso')[:5]
    
    # Vehículos listos para retiro (estado 'listo')
    vehiculos_listos = Vehiculo.objects.filter(estado='listo').order_by('-fecha_ingreso')[:5]
    
    context = {
        'vehiculos_ingresados': vehiculos_ingresados,
        'ultimos_ingresos': ultimos_ingresos,
        'vehiculos_taller': vehiculos_taller,
        'vehiculos_listos': vehiculos_listos,
        'estadisticas': estadisticas,
        'hoy': hoy,
    }
    
    return render(request, 'usuarios/dashboard_guardia.html', context)

# usuarios/views.py - AGREGAR
@login_required
def dashboard_ehs(request):
    """Dashboard específico para EHS - Solo seguridad y salud"""
    if not (request.user.rol == 'ehs' or request.user.rol == 'admin' or request.user.is_superuser):
        messages.error(request, 'No tienes acceso a esta vista.')
        return redirect('clinica_dashboard')
    
    # Tareas de alto riesgo
    tareas_alto_riesgo = Tarea.objects.filter(prioridad='alta').select_related(
        'vehiculo', 'mecanico_asignado'
    )
    
    # Siniestros activos (si existe el modelo)
    try:
        from vehiculos.models import Siniestro
        siniestros_activos = Siniestro.objects.filter(
            estado__in=['reportado', 'en_evaluacion']
        ).select_related('vehiculo')
    except:
        siniestros_activos = []
    
    # Pausas por problemas de seguridad (ejemplo)
    pausas_seguridad = Pausa.objects.filter(
        motivo__in=['problema_calidad', 'consultas_tecnicas']
    ).select_related('tarea', 'tarea__vehiculo')[:10]
    
    # Estadísticas de seguridad
    estadisticas = {
        'tareas_alto_riesgo': tareas_alto_riesgo.count(),
        'siniestros_activos': len(siniestros_activos),
        'pausas_seguridad': pausas_seguridad.count(),
        'vehiculos_taller': Vehiculo.objects.exclude(estado='entregado').count(),
    }
    
    context = {
        'tareas_alto_riesgo': tareas_alto_riesgo,
        'siniestros_activos': siniestros_activos,
        'pausas_seguridad': pausas_seguridad,
        'estadisticas': estadisticas,
    }
    
    return render(request, 'usuarios/dashboard_ehs.html', context)


@login_required
def dashboard_recepcionista(request):
    """Dashboard para Recepcionista de Vehículos"""
    # Permitir acceso a recepcionistas, admins y superusers
    if not (request.user.rol == 'recepcionista' or request.user.rol == 'admin' or request.user.is_superuser):
        messages.error(request, 'No tienes acceso a esta vista.')
        return redirect('clinica_dashboard')
    
    # Vehículos ingresados hoy
    hoy = timezone.now().date()
    vehiculos_hoy = Vehiculo.objects.filter(
        fecha_ingreso__date=hoy
    ).select_related('chofer_asignado')
    
    # Citas programadas para hoy
    from agenda.models import CitaMantenimiento
    citas_hoy = CitaMantenimiento.objects.filter(
        fecha_hora__date=hoy,
        estado__in=['pendiente', 'confirmada']
    ).select_related('vehiculo', 'solicitante')
    
    # Vehículos en taller
    vehiculos_en_taller = Vehiculo.objects.exclude(
        estado='entregado'
    ).select_related('chofer_asignado')
    
    # Tareas recientes
    tareas_recientes = Tarea.objects.all().select_related(
        'vehiculo', 'mecanico_asignado'
    ).order_by('-fecha_creacion')[:10]
    
    # Estadísticas
    estadisticas = {
        'vehiculos_hoy': vehiculos_hoy.count(),
        'citas_hoy': citas_hoy.count(),
        'vehiculos_taller': vehiculos_en_taller.count(),
        'tareas_pendientes': Tarea.objects.filter(estado='pendiente').count(),
    }
    
    context = {
        'vehiculos_hoy': vehiculos_hoy,
        'citas_hoy': citas_hoy,
        'vehiculos_en_taller': vehiculos_en_taller,
        'tareas_recientes': tareas_recientes,
        'estadisticas': estadisticas,
    }
    return render(request, 'usuarios/dashboard_recepcionista.html', context)


@login_required
def dashboard_vendedor(request):
    """Dashboard específico para vendedores/preventistas"""
    if not (request.user.rol == 'vendedor' or request.user.rol == 'admin' or request.user.is_superuser):
        messages.error(request, 'No tienes acceso a esta vista.')
        return redirect('clinica_dashboard')
    
    # Vehículos asignados al vendedor (usando chofer_asignado)
    vehiculos = Vehiculo.objects.filter(chofer_asignado=request.user).prefetch_related('tareas')
    
    # Citas del vendedor (por patente o como solicitante)
    mis_patentes = vehiculos.values_list('patente', flat=True)
    citas = CitaMantenimiento.objects.filter(
        models.Q(solicitante=request.user) | 
        models.Q(patente__in=mis_patentes)
    ).select_related('vehiculo').order_by('fecha_hora')
    
    # Fechas clave
    from django.utils import timezone
    from datetime import timedelta
    
    hoy = timezone.now().date()
    treinta_dias = hoy + timedelta(days=30)
    
    # Documentos próximos a vencer
    documentos_proximos = DocumentoVehiculo.objects.filter(
        vehiculo__in=vehiculos,
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=treinta_dias
    ).select_related('vehiculo')
    
    # Documentos vencidos
    documentos_vencidos = DocumentoVehiculo.objects.filter(
        vehiculo__in=vehiculos,
        fecha_vencimiento__lt=hoy
    ).select_related('vehiculo')
    
    # Citas próximas (7 días)
    citas_proximas = citas.filter(
        fecha_hora__date__gte=hoy,
        fecha_hora__date__lte=hoy + timedelta(days=7),
        estado__in=['pendiente', 'confirmada']
    )
    
    # ✅ Estadísticas COMPLETAS (incluyendo en_taller y listos_retiro)
    en_taller = vehiculos.exclude(estado='entregado').count()
    listos_retiro = vehiculos.filter(estado='listo').count()
    
    estadisticas = {
        'total_vehiculos': vehiculos.count(),
        'citas_pendientes': citas.filter(estado='pendiente').count(),
        'citas_confirmadas': citas.filter(estado='confirmada').count(),
        'documentos_proximos': documentos_proximos.count(),
        'documentos_vencidos': documentos_vencidos.count(),
        'citas_proximas': citas_proximas.count(),
        # ✅ Campos que tu HTML necesita:
        'en_taller': en_taller,
        'listos_retiro': listos_retiro,
    }
    
    context = {
        'vehiculos': vehiculos,
        'citas': citas,
        'citas_proximas': citas_proximas,
        'documentos_proximos': documentos_proximos,
        'documentos_vencidos': documentos_vencidos,
        'estadisticas': estadisticas,
    }
    
    return render(request, 'usuarios/dashboard_vendedor.html', context)

@login_required
def dashboard_bodeguero(request):
    """Dashboard específico para el bodeguero: gestión de inventario y movimientos."""
    if not (request.user.rol == 'bodeguero' or request.user.rol == 'admin' or request.user.is_superuser):
        messages.error(request, 'No tienes acceso a esta vista.')
        return redirect('clinica_dashboard')

    hoy = timezone.now().date()

    # 1. Métricas de inventario
    total_repuestos = Repuesto.objects.count()
    repuestos_bajo_stock = Repuesto.objects.filter(estado='bajo_stock').count()
    repuestos_agotados = Repuesto.objects.filter(estado='agotado').count()

    # 2. Solicitudes activas: tareas en progreso o pendientes
    #   (asumimos que toda tarea activa requiere repuestos)
    solicitudes_pendientes = Tarea.objects.filter(
        estado__in=['pendiente', 'en_proceso']
    ).select_related('vehiculo', 'mecanico_asignado')

    # 3. Movimientos de hoy
    movimientos_hoy = MovimientoInventario.objects.filter(
        fecha_movimiento__date=hoy
    ).select_related('repuesto', 'usuario', 'tarea')

    # 4. Stock crítico
    repuestos_criticos = Repuesto.objects.filter(
        estado__in=['bajo_stock', 'agotado']
    ).order_by('stock_actual')[:10]

    context = {
        'estadisticas': {
            'total_repuestos': total_repuestos,
            'repuestos_bajo_stock': repuestos_bajo_stock,
            'repuestos_agotados': repuestos_agotados,
            'solicitudes_pendientes': solicitudes_pendientes.count(),
            'movimientos_hoy': movimientos_hoy.count(),
        },
        'solicitudes_pendientes': solicitudes_pendientes,
        'repuestos_criticos': repuestos_criticos,
        'ultimos_movimientos': movimientos_hoy.order_by('-fecha_movimiento')[:8],
    }

    return render(request, 'usuarios/dashboard_bodeguero.html', context)
@login_required
def dashboard_admin(request):
    """Dashboard exclusivo para el Administrador: Gestión de usuarios y acceso global."""
    # Verificar rol de administrador
    if request.user.rol != 'admin':
        messages.error(request, 'No tienes permisos para acceder a esta vista.')
        # Opcional: redirigir a un dashboard genérico si es que existe para otros roles no autorizados
        # return redirect('clinica_dashboard') # Asegúrate que esta URL exista y sea segura
        return redirect('login') # O a una página de error 403 personalizada

    # Métricas Generales (opcional para el admin)
    total_usuarios = Usuario.objects.count()
    total_vehiculos = Vehiculo.objects.count()
    total_tareas = Tarea.objects.count()
    total_repuestos = Repuesto.objects.count()

    context = {
        'estadisticas': {
            'total_usuarios': total_usuarios,
            'total_vehiculos': total_vehiculos,
            'total_tareas': total_tareas,
            'total_repuestos': total_repuestos,
        },
        # Opcional: puedes pasar listas resumidas aquí también
        # 'ultimos_usuarios': Usuario.objects.all().order_by('-date_joined')[:5],
        # 'ultimas_tareas': Tarea.objects.all().order_by('-fecha_creacion')[:5],
    }

    return render(request, 'usuarios/dashboard_admin.html', context)

@login_required
def dashboard_jefe_taller(request):
    """Dashboard exclusivo para el Jefe de Taller: supervisión, asignación y productividad."""
    if not (request.user.rol == 'jefe_taller' or request.user.rol == 'admin' or request.user.is_superuser):
        messages.error(request, 'No tienes acceso a esta vista.')
        return redirect('clinica_dashboard')

    # 1. Vehículos en taller
    vehiculos_en_taller = Vehiculo.objects.exclude(estado='entregado').select_related('mecanico_asignado', 'chofer_asignado')

    # 2. Tareas: pendientes (sin mecánico) y activas (con mecánico)
    tareas_pendientes = Tarea.objects.filter(estado='pendiente', mecanico_asignado__isnull=True).select_related('vehiculo')
    tareas_activas = Tarea.objects.filter(
        estado__in=['en_proceso', 'pausada']
    ).select_related('vehiculo', 'mecanico_asignado')

    # 3. Mecánicos activos (para asignar)
    mecanicos = Usuario.objects.filter(rol='mecanico', is_active=True)

    # 4. Citas pendientes
    from agenda.models import CitaMantenimiento
    citas_pendientes = CitaMantenimiento.objects.filter(estado='pendiente').select_related('vehiculo', 'solicitante').order_by('fecha_hora')

    # 5. Productividad por mecánico (últimos 30 días)
    from django.utils import timezone
    hace_30_dias = timezone.now() - timezone.timedelta(days=30)
    
    productividad = []
    for mecanico in mecanicos:
        tareas_completadas = Tarea.objects.filter(
            mecanico_asignado=mecanico,
            estado='completada',
            fecha_fin__gte=hace_30_dias
        ).count()
        tareas_totales = Tarea.objects.filter(
            mecanico_asignado=mecanico,
            fecha_creacion__gte=hace_30_dias
        ).count()
        eficiencia = round((tareas_completadas / tareas_totales * 100), 1) if tareas_totales > 0 else 0
        productividad.append({
            'mecanico': mecanico,
            'completadas': tareas_completadas,
            'totales': tareas_totales,
            'eficiencia': eficiencia
        })

    # 6. Estadísticas generales
    estadisticas = {
        'total_vehiculos': vehiculos_en_taller.count(),
        'tareas_pendientes': tareas_pendientes.count(),
        'tareas_activas': tareas_activas.count(),
        'mecanicos_activos': mecanicos.count(),
        'listos_para_retiro': Vehiculo.objects.filter(estado='listo').count(),
        'citas_pendientes': citas_pendientes.count(),
    }

    context = {
        'vehiculos_en_taller': vehiculos_en_taller,
        'tareas_pendientes': tareas_pendientes,
        'tareas_activas': tareas_activas,
        'mecanicos': mecanicos,
        'productividad': productividad,
        'estadisticas': estadisticas,
        'citas_pendientes': citas_pendientes,
    }
    return render(request, 'usuarios/dashboard_jefe_taller.html', context)

# usuarios/views.py
@login_required
def asignar_tarea(request, tarea_id):
    """Asignar un mecánico a una tarea pendiente."""
    if not (request.user.rol == 'jefe_taller' or request.user.rol == 'admin' or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para asignar tareas.')
        return redirect('dashboard_jefe_taller')

    tarea = get_object_or_404(Tarea, id=tarea_id, estado='pendiente', mecanico_asignado__isnull=True)
    
    if request.method == 'POST':
        mecanico_id = request.POST.get('mecanico_id')
        if mecanico_id:
            try:
                mecanico = Usuario.objects.get(id=mecanico_id, rol='mecanico')
                tarea.mecanico_asignado = mecanico
                tarea.estado = 'en_proceso'
                tarea.fecha_inicio = timezone.now()
                tarea.save()
                messages.success(request, f'Tarea asignada a {mecanico.get_full_name()}')
            except Usuario.DoesNotExist:
                messages.error(request, 'Mecánico no válido.')
        else:
            messages.error(request, 'Selecciona un mecánico.')
    
    return redirect('dashboard_jefe_taller')
@login_required
def importar_usuarios_excel(request):
    if not (request.user.is_superuser or getattr(request.user, 'rol', '') == 'admin'):
        messages.error(request, 'No tienes permiso para importar usuarios.')
        return redirect('lista_usuarios')

    if request.method != 'POST':
        return redirect('lista_usuarios')

    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Debes seleccionar un archivo .xlsx o .csv')
        return redirect('lista_usuarios')

    created = 0
    updated = 0
    errors = []

    def parse_bool(val):
        s = str(val or '').strip().upper()
        return s in ['SI', 'SÍ', 'YES', 'TRUE', '1']

    def normalize_role(val):
        s = str(val or '').strip()
        allowed = [c for c, _ in Usuario.ROLES]
        return s if s in allowed else None

    ext = archivo.name.lower().split('.')[-1]
    rows = []
    headers = []

    try:
        if ext in ['xlsx', 'xls']:
            from openpyxl import load_workbook
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h or '').strip().lower() for h in row]
                else:
                    rows.append([c for c in row])
        elif ext == 'csv':
            import io, csv
            content = archivo.read().decode('utf-8-sig')
            delimiter = ';' if ';' in content.splitlines()[0] else ','
            reader = csv.reader(io.StringIO(content), delimiter=delimiter)
            for i, row in enumerate(reader):
                if i == 0:
                    headers = [str(h or '').strip().lower() for h in row]
                else:
                    rows.append(row)
        else:
            messages.error(request, 'Formato no soportado. Usa .xlsx o .csv')
            return redirect('lista_usuarios')
    except Exception as e:
        messages.error(request, f'Error al leer el archivo: {e}')
        return redirect('lista_usuarios')

    index = {h: i for i, h in enumerate(headers)}
    required = ['username', 'email', 'first_name', 'last_name', 'rol']
    missing = [h for h in required if h not in index]
    if missing:
        messages.error(request, f'Faltan columnas requeridas: {", ".join(missing)}')
        return redirect('lista_usuarios')

    for row in rows:
        try:
            username = str(row[index['username']]).strip()
            email = str(row[index['email']]).strip()
            first_name = str(row[index['first_name']]).strip()
            last_name = str(row[index['last_name']]).strip()
            rol_val = normalize_role(row[index['rol']])
            telefono = str(row[index.get('telefono', None)]).strip() if index.get('telefono') is not None else ''
            activo = parse_bool(row[index.get('activo', None)]) if index.get('activo') is not None else True
            is_superuser = parse_bool(row[index.get('is_superuser', None)]) if index.get('is_superuser') is not None else False
            password = str(row[index.get('password', None)]).strip() if index.get('password') is not None else ''

            if not rol_val:
                errors.append(f'Usuario {username}: rol inválido')
                continue
            if not username:
                errors.append('Fila sin username')
                continue

            usuario = Usuario.objects.filter(username=username).first()
            if usuario:
                usuario.email = email
                usuario.first_name = first_name
                usuario.last_name = last_name
                usuario.rol = rol_val
                usuario.telefono = telefono or usuario.telefono
                usuario.is_active = activo
                usuario.is_superuser = is_superuser
                if password:
                    usuario.set_password(password)
                usuario.save()
                updated += 1
            else:
                usuario = Usuario(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    rol=rol_val,
                    telefono=telefono,
                    is_active=activo,
                    is_superuser=is_superuser,
                )
                usuario.set_password(password or 'Tempor@l123')
                usuario.save()
                created += 1
        except Exception as e:
            errors.append(str(e))

    if created or updated:
        messages.success(request, f'{created} creados, {updated} actualizados')
    if errors:
        messages.warning(request, f'Errores: {len(errors)}')

    return redirect('lista_usuarios')
@login_required
def importar_usuarios(request):
    if not (request.user.is_superuser or request.user.rol == 'admin'):
        messages.error(request, 'No tienes permiso para importar usuarios.')
        return redirect('lista_usuarios')
    ejemplo_csv = (
        'username,email,first_name,last_name,rol,telefono,activo,is_superuser,password\n'
        'jtaller,jefe@empresa.cl,Juan,Taller,jefe_taller,+56911111111,SI,NO,Tempor@l123\n'
        'mmecanico1,mec1@empresa.cl,Marta,Mecanico,mecanico,+56922222222,SI,NO,Tempor@l123\n'
        'gseguridad,guardia@empresa.cl,Carlos,Guardia,guardia,+56933333333,SI,NO,Tempor@l123\n'
        'adminchile,admin@empresa.cl,Ana,Admin,admin,+56944444444,SI,SI,Admin$2025\n'
        'chofer01,chofer01@empresa.cl,Pedro,Muñoz,chofer,+56955555555,SI,NO,Tempor@l123'
    )
    return render(request, 'usuarios/importar_usuarios.html', {
        'ejemplo_csv': ejemplo_csv,
    })
